"""
Part 5 — Create Excel output with five sheets.

Sheet 1 "Summary"                   – Key metrics per company
Sheet 2 "Company Info"              – All raw CVR data
Sheet 3 "Financial Data"            – Structured annual report figures per year
Sheet 4 "Financial Items"           – All raw financial XBRL elements per year
Sheet 5 "Miscellaneous report items"– Non-financial XBRL elements per year
"""
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── Colour palette ─────────────────────────────────────────────────────────────
_DARK_BLUE  = PatternFill("solid", fgColor="1F4E79")
_LIGHT_BLUE = PatternFill("solid", fgColor="D6E4F0")
_WHITE      = PatternFill("solid", fgColor="FFFFFF")
_GREEN      = PatternFill("solid", fgColor="C6EFCE")
_RED        = PatternFill("solid", fgColor="FFC7CE")
_ORANGE     = PatternFill("solid", fgColor="FCE4D6")
_GREY       = PatternFill("solid", fgColor="D9D9D9")

_HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
_NORM_FONT = Font(size=10)

_FMT_INT   = '#,##0'
_FMT_DEC1  = '#,##0.0'
_FMT_DEC2  = '#,##0.00'
_FMT_PCT   = '0.0"%"'

_NOT_FOUND = "item not found"


# ── Styling helpers ────────────────────────────────────────────────────────────

def _s(cell, value=None, font=None, fill=None, align="left", fmt=None):
    if value is not None:
        try:
            cell.value = value
        except Exception:
            cell.value = "error"
    cell.font      = font or _NORM_FONT
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt


def _header_row(ws, headers: list):
    for col, h in enumerate(headers, 1):
        _s(ws.cell(row=1, column=col), value=h,
           font=_HDR_FONT, fill=_DARK_BLUE, align="center")
    ws.row_dimensions[1].height = 22


def _widths(ws, w: list):
    for col, width in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _zebra(row: int):
    return _LIGHT_BLUE if row % 2 == 0 else _WHITE


def _join(lst) -> str:
    return ", ".join(lst) if lst else ""


def _rev(reports: list, idx: int, field: str):
    return reports[idx].get(field) if idx < len(reports) else None


def _safe_div(a, b, pct=False):
    if a is not None and b and b != 0:
        return (a / b * 100) if pct else (a / b)
    return None


def _determine_currency(companies: list) -> str:
    currencies = []
    for c in companies:
        for rep in c.get("reports", []):
            cur = rep.get("currency")
            if cur:
                currencies.append(cur)
    if not currencies:
        return "DKK"
    return Counter(currencies).most_common(1)[0][0]


# ── Sheet 1: Summary ───────────────────────────────────────────────────────────

def _build_summary_cols(cur: str) -> tuple:
    """Returns (headers, dkk_idx_set, pct_idx_set, bool_idx_set, int_idx_set) — 1-based."""
    c = cur
    headers = [
        # Identifikation (frozen)
        "CVR", "Company",
        # Virksomhedsprofil
        "City", "Region", "Industry", "No. of reports fetched",
        # Medarbejdere
        "Employees (latest)",
        "No. Employees (latest)", "No. Employees -1y", "No. Employees -2y", "No. Employees -3y",
        # Revenue
        f"Revenue (latest, {c})", f"Revenue -1y ({c})", f"Revenue -2y ({c})", f"Revenue -3y ({c})",
        "Revenue CAGR (%)", "Revenue CAGR period", "Positive revenue trend?",
        # EBITDA
        f"EBITDA (latest, {c})", f"EBITDA -1y ({c})", f"EBITDA -2y ({c})", f"EBITDA -3y ({c})",
        "EBITDA CAGR (%)", "EBITDA CAGR period", "Positive EBITDA Trend", "EBITDA trend period",
        # EBIT
        f"EBIT (latest, {c})", f"EBIT -1y ({c})", f"EBIT -2y ({c})", f"EBIT -3y ({c})",
        "EBIT CAGR (%)", "EBIT CAGR period", "Positive EBIT Trend", "EBIT trend period",
        # Øvrig økonomi
        f"Net profit (latest, {c})", "Rule-of-40",
        # Kontakt & outreach
        "Board members", "Contact person", "Contact person role",
        "Website", "LinkedIn", "Comments",
    ]
    #                      1   2   3   4   5   6   7   8   9  10  11
    #                     12  13  14  15  16  17  18
    #                     19  20  21  22  23  24  25  26
    #                     27  28  29  30  31  32  33  34
    #                     35  36  37  38  39  40  41  42
    dkk_cols  = {12, 13, 14, 15, 19, 20, 21, 22, 27, 28, 29, 30, 35}
    pct_cols  = {16, 23, 31, 36}
    bool_cols = {18, 25, 33}
    int_cols  = {8, 9, 10, 11}
    return headers, dkk_cols, pct_cols, bool_cols, int_cols


def _summary_row(company: dict, cur: str) -> list:
    r   = company.get("reports", [])
    kpi = company.get("kpi", {})
    inp = company.get("_input", {})

    return [
        # Identifikation
        str(company.get("cvr", "")),
        company.get("name", ""),
        # Virksomhedsprofil
        company.get("city", ""),
        company.get("region", ""),
        company.get("industry", ""),
        len(r),
        # Medarbejdere
        company.get("employees", ""),
        _rev(r, 0, "employees_xbrl"), _rev(r, 1, "employees_xbrl"),
        _rev(r, 2, "employees_xbrl"), _rev(r, 3, "employees_xbrl"),
        # Revenue
        _rev(r, 0, "revenue"), _rev(r, 1, "revenue"),
        _rev(r, 2, "revenue"), _rev(r, 3, "revenue"),
        kpi.get("revenue_cagr"), kpi.get("revenue_cagr_span"),
        kpi.get("positive_revenue_trend"),
        # EBITDA
        _rev(r, 0, "ebitda"), _rev(r, 1, "ebitda"),
        _rev(r, 2, "ebitda"), _rev(r, 3, "ebitda"),
        kpi.get("ebitda_cagr"), kpi.get("ebitda_cagr_span"),
        kpi.get("positive_ebitda_trend"), kpi.get("ebitda_trend_period"),
        # EBIT
        _rev(r, 0, "ebit"), _rev(r, 1, "ebit"),
        _rev(r, 2, "ebit"), _rev(r, 3, "ebit"),
        kpi.get("ebit_cagr"), kpi.get("ebit_cagr_span"),
        kpi.get("positive_ebit_trend"), kpi.get("ebit_trend_period"),
        # Øvrig økonomi
        _rev(r, 0, "net_profit"),
        kpi.get("rule_of_40"),
        # Kontakt & outreach
        _join(company.get("board_members", [])),
        inp.get("Contact person", ""),
        inp.get("Contact person role", ""),
        company.get("website", "") or inp.get("Website", ""),
        inp.get("LinkedIn", ""),
        _da_comment(r),
    ]


def _da_comment(reports: list) -> str:
    if not reports:
        return ""
    latest = reports[0]
    if latest.get("depreciation_is_fsa") is False:
        element = latest.get("depreciation_elements", "")
        return (
            f"D&A for latest report based on non-FSA element '{element}' — "
            f"see Financial Data sheet, column 'Depreciation and Amortisation elements used'"
        )
    return ""


def create_summary_sheet(ws, companies: list, extra_cols: list, cur: str):
    ws.title = "Summary"

    base_headers, dkk_cols, pct_cols, bool_cols, int_cols = _build_summary_cols(cur)
    extra = [c for c in extra_cols if c not in base_headers]
    headers = base_headers + extra

    _header_row(ws, headers)

    for row, company in enumerate(companies, 2):
        zebra = _zebra(row)
        vals  = _summary_row(company, cur)
        inp   = company.get("_input", {})

        for col_name in extra:
            vals.append(inp.get(col_name, ""))

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col)
            if val == "N/A":
                _s(cell, value="N/A", fill=_GREY, align="center")
            elif col in bool_cols and val is not None:
                fill = _GREEN if val else _RED
                _s(cell, value="TRUE" if val else "FALSE", fill=fill, align="center")
            elif col in dkk_cols:
                _s(cell, value=val, fill=zebra, align="right", fmt=_FMT_INT)
            elif col in pct_cols:
                _s(cell, value=val, fill=zebra, align="right", fmt=_FMT_DEC1)
            elif col in int_cols:
                _s(cell, value=int(val) if val is not None else None, fill=zebra, align="right", fmt=_FMT_INT)
            else:
                _s(cell, value=val, fill=zebra)

    ws.freeze_panes = "C2"
    base_w = [
        12, 28,               # CVR, Company
        16, 22, 30, 10,       # City, Region, Industry, No. reports
        14, 13, 13, 13, 13,   # Employees, No.Emp x4
        18, 16, 16, 16,       # Revenue x4
        14, 22, 14,           # Revenue CAGR, period, trend
        18, 16, 16, 16,       # EBITDA x4
        14, 22, 14, 22,       # EBITDA CAGR, period, trend, trend period
        18, 16, 16, 16,       # EBIT x4
        14, 22, 14, 22,       # EBIT CAGR, period, trend, trend period
        18, 12,               # Net profit, Rule-of-40
        32, 20, 20,           # Board members, Contact x2
        28, 28, 60,           # Website, LinkedIn, Comments
    ]
    _widths(ws, base_w + [16] * len(extra))


# ── Sheet 2: Company Info ──────────────────────────────────────────────────────

_COMPANY_COLS = [
    "CVR", "Company name", "Status", "Company type", "Founded",
    "Postal code", "City", "Municipality", "Region",
    "Industry", "Industry code",
    "Employees", "Employees year",
    "Phone", "Email", "Website",
    "Directors", "Board members", "Owners",
    "Searched via", "Matches found",
]


def create_company_sheet(ws, companies: list):
    ws.title = "Company Info"
    _header_row(ws, _COMPANY_COLS)

    for row, c in enumerate(companies, 2):
        zebra = _zebra(row)
        vals  = [
            str(c.get("cvr", "")),
            c.get("name", ""),
            c.get("status", ""),
            c.get("virksomhedsform", ""),
            c.get("founded", ""),
            c.get("postal_code", ""),
            c.get("city", ""),
            c.get("municipality", ""),
            c.get("region", ""),
            c.get("industry", ""),
            c.get("industry_code", ""),
            c.get("employees", ""),
            c.get("employees_year", ""),
            c.get("phone", ""),
            c.get("email", ""),
            c.get("website", ""),
            _join(c.get("directors", [])),
            _join(c.get("board_members", [])),
            _join(c.get("owners", [])),
            c.get("_soegt_via", ""),
            c.get("_antal_matches", ""),
        ]
        for col, val in enumerate(vals, 1):
            _s(ws.cell(row=row, column=col), value=val, fill=zebra)

    ws.freeze_panes = "C2"
    _widths(ws, [12, 32, 12, 14, 12, 10, 16, 18, 22,
                 32, 10, 12, 12, 14, 24, 28,
                 28, 28, 28, 14, 12])


# ── Sheet 3: Financial Data ────────────────────────────────────────────────────

def _build_fin_spec(cur: str) -> list:
    """
    Returns list of (header, field_or_None, fmt) tuples.
    field=None means the value is calculated via _CALC_MAP.
    fmt: 'int' | 'pct' | 'ratio' | 'text'
    """
    c = cur
    return [
        # Identifikation (frozen)
        ("CVR",                                        "cvr_str",                    "text"),
        ("Company",                                    "name",                       "text"),
        ("Year",                                       "year",                       "text"),
        ("Period end",                                 "period_end",                 "text"),
        ("Report format",                              "filing_type",                "text"),
        # Datakvalitet
        ("Reporting class",                            "reporting_class",            "text"),
        ("Taxonomy version",                           "taxonomy_version",           "text"),
        # Resultatopgørelse — omsætning
        (f"Revenue ({c})",                             "revenue",                    "int"),
        (f"Cost of sales ({c})",                       "cost_of_sales",              "int"),
        (f"Gross profit ({c})",                        "gross_profit",               "int"),
        # Driftsomkostninger
        (f"External expenses ({c})",                   "external_expenses",          "int"),
        (f"Employee benefit expenses ({c})",           "employee_benefits",          "int"),
        (f"Wages and salaries ({c})",                  "wages_salaries",             "int"),
        (f"Other operating income ({c})",              "other_operating_income",     "int"),
        ("No. of employees",                           "employees_xbrl",             "int"),
        # Driftsresultat
        (f"EBITDA ({c})",                              "ebitda",                     "int"),
        ("EBITDA margin (%)",                          None,                         "pct"),
        (f"Depreciation and amortisation ({c})",       "depreciation",               "int"),
        ("No. of D&A elements used",                   "depreciation_element_count", "int"),
        ("D&A elements used",                          "depreciation_elements",      "text"),
        (f"EBIT ({c})",                                "ebit",                       "int"),
        ("EBIT margin (%)",                            None,                         "pct"),
        ("No. of EBIT elements used",                  "ebit_elements_count",        "int"),
        ("EBIT elements used",                         "ebit_elements",              "text"),
        # Finansielle poster og skat
        (f"Finance income ({c})",                      "finance_income",             "int"),
        (f"Finance expenses ({c})",                    "finance_expenses",           "int"),
        (f"Profit before tax ({c})",                   "profit_before_tax",          "int"),
        (f"Tax expense ({c})",                         "tax_expense",                "int"),
        (f"Current tax expense ({c})",                 "current_tax",                "int"),
        (f"Deferred tax expense ({c})",                "deferred_tax",               "int"),
        (f"Net profit ({c})",                          "net_profit",                 "int"),
        ("Net profit margin (%)",                      None,                         "pct"),
        # Balance — Aktiver
        (f"Total assets ({c})",                        "assets",                     "int"),
        (f"Noncurrent assets ({c})",                   "noncurrent_assets",          "int"),
        (f"Current assets ({c})",                      "current_assets",             "int"),
        (f"Cash and cash equivalents ({c})",           "cash",                       "int"),
        (f"Shortterm receivables ({c})",               "shortterm_receivables",      "int"),
        # Balance — Passiver
        (f"Equity ({c})",                              "equity",                     "int"),
        (f"Contributed capital ({c})",                 "contributed_capital",        "int"),
        (f"Retained earnings ({c})",                   "retained_earnings",          "int"),
        (f"Proposed dividend ({c})",                   "proposed_dividend",          "int"),
        (f"Longterm liabilities ({c})",                "longterm_liabilities",       "int"),
        (f"Shortterm liabilities ({c})",               "shortterm_liabilities",      "int"),
        (f"Provisions ({c})",                          "provisions",                 "int"),
        # Nøgletal
        ("Equity ratio (%)",                           None,                         "pct"),
        ("Current ratio",                              None,                         "ratio"),
        ("Return on equity (%)",                       None,                         "pct"),
    ]


def _calc_fin_row(reg: dict, cvr: str, name: str) -> dict:
    """Pre-calculate all values for a Financial Data row."""
    rev  = reg.get("revenue")
    eb   = reg.get("ebitda")
    eq   = reg.get("equity")
    as_  = reg.get("assets")
    ca   = reg.get("current_assets")
    sl   = reg.get("shortterm_liabilities")
    ebit = reg.get("ebit")
    np_  = reg.get("net_profit")

    return {
        "cvr_str":                   cvr,
        "name":                      name,
        "year":                      reg.get("year", ""),
        "period_end":                reg.get("period_end", ""),
        "reporting_class":           reg.get("reporting_class", ""),
        "revenue":                   rev,
        "gross_profit":              reg.get("gross_profit"),
        "ebit":                      ebit,
        "ebit_elements":             reg.get("ebit_elements", ""),
        "ebit_elements_count":       1 if reg.get("ebit_elements") else None,
        "depreciation":              reg.get("depreciation"),
        "ebitda":                    eb,
        "_ebitda_margin":            _safe_div(eb, rev, pct=True),
        "net_profit":                np_,
        "employees_xbrl":            int(reg["employees_xbrl"]) if reg.get("employees_xbrl") is not None else None,
        "cost_of_sales":             reg.get("cost_of_sales"),
        "other_operating_income":    reg.get("other_operating_income"),
        "external_expenses":         reg.get("external_expenses"),
        "employee_benefits":         reg.get("employee_benefits"),
        "wages_salaries":            reg.get("wages_salaries"),
        "tax_expense":               reg.get("tax_expense"),
        "current_tax":               reg.get("current_tax"),
        "deferred_tax":              reg.get("deferred_tax"),
        "finance_income":            reg.get("finance_income"),
        "finance_expenses":          reg.get("finance_expenses"),
        "profit_before_tax":         reg.get("profit_before_tax"),
        "assets":                    as_,
        "equity":                    eq,
        "current_assets":            ca,
        "noncurrent_assets":         reg.get("noncurrent_assets"),
        "cash":                      reg.get("cash"),
        "shortterm_receivables":     reg.get("shortterm_receivables"),
        "shortterm_liabilities":     sl,
        "longterm_liabilities":      reg.get("longterm_liabilities"),
        "provisions":                reg.get("provisions"),
        "contributed_capital":       reg.get("contributed_capital"),
        "retained_earnings":         reg.get("retained_earnings"),
        "proposed_dividend":         reg.get("proposed_dividend"),
        "_equity_ratio":             _safe_div(eq, as_, pct=True),
        "_current_ratio":            _safe_div(ca, sl),
        "_roe":                      _safe_div(np_, eq, pct=True),
        "_ebit_margin":              _safe_div(ebit, rev, pct=True),
        "_np_margin":                _safe_div(np_, rev, pct=True),
        "taxonomy_version":          reg.get("taxonomy_version", ""),
        "depreciation_element_count": reg.get("depreciation_element_count"),
        "depreciation_elements":     reg.get("depreciation_elements", ""),
    }


_CALC_MAP = {
    "EBITDA margin (%)":    "_ebitda_margin",
    "Equity ratio (%)":     "_equity_ratio",
    "Current ratio":        "_current_ratio",
    "Return on equity (%)": "_roe",
    "EBIT margin (%)":      "_ebit_margin",
    "Net profit margin (%)":"_np_margin",
}


def create_financial_sheet(ws, companies: list, cur: str):
    ws.title = "Financial Data"

    spec    = _build_fin_spec(cur)
    headers = [s[0] for s in spec]
    _header_row(ws, headers)

    row = 2
    for c in companies:
        cvr  = str(c.get("cvr", ""))
        name = c.get("name", "")

        for reg in c.get("reports", []):
            zebra = _zebra(row)
            vals  = _calc_fin_row(reg, cvr, name)
            vals["filing_type"] = c.get("filing_type", "XBRL")

            for col, (header, field, fmt) in enumerate(spec, 1):
                cell = ws.cell(row=row, column=col)
                val  = vals.get(_CALC_MAP.get(header)) if field is None else vals.get(field)

                if fmt == "int":
                    _s(cell, value=val, fill=zebra, align="right", fmt=_FMT_INT)
                elif fmt == "pct":
                    _s(cell, value=val, fill=zebra, align="right", fmt=_FMT_DEC1)
                elif fmt == "ratio":
                    _s(cell, value=val, fill=zebra, align="right", fmt=_FMT_DEC2)
                else:
                    _s(cell, value=val, fill=zebra)

            row += 1

    ws.freeze_panes = "F2"
    widths = [
        12, 30, 8, 14, 14,   # CVR, Company, Year, Period end, Report format
        18, 28,              # Reporting class, Taxonomy version
        18, 18, 18,          # Revenue, Cost of sales, Gross profit
        18, 18, 18, 18, 14,  # External exp, Employee benefits, Wages, Other op. income, Employees
        18, 14,              # EBITDA, EBITDA margin
        18, 10, 50,          # D&A, D&A count, D&A elements
        18, 14,              # EBIT, EBIT margin
        10, 50,              # EBIT count, EBIT elements
        18, 18, 18,          # Finance income, Finance expenses, Profit before tax
        18, 18, 18,          # Tax, Current tax, Deferred tax
        18, 14,              # Net profit, Net profit margin
        18, 18, 18, 18, 18,  # Total assets, Noncurrent, Current, Cash, ST receivables
        18, 18, 18, 18,      # Equity, Contributed capital, Retained earnings, Proposed dividend
        18, 18, 18,          # LT liabilities, ST liabilities, Provisions
        14, 14, 14,          # Equity ratio, Current ratio, ROE
    ]
    _widths(ws, widths)


# ── Namespace → short prefix mapping ──────────────────────────────────────────

_NS_SHORT = {
    "http://xbrl.dcca.dk/fsa": "fsa",
    "http://xbrl.dcca.dk/cmn": "cmn",
    "http://xbrl.dcca.dk/gsd": "gsd",
    "http://xbrl.dcca.dk/sob": "sob",
    "http://xbrl.dcca.dk/arr": "arr",
    "http://xbrl.dcca.dk/mrv": "mrv",
}

def _ns_prefix(ns_uri: str) -> str:
    if ns_uri in _NS_SHORT:
        return _NS_SHORT[ns_uri]
    if "ifrs" in ns_uri.lower():
        return "ifrs"
    if ns_uri:
        return ns_uri.rstrip("/").split("/")[-1][:10]
    return "?"


# ── Sheet 4: Financial Items ───────────────────────────────────────────────────

def _collect_financial_keys(companies: list) -> list:
    """Returns sorted list of (ns_uri, local_name) tuples across all reports."""
    seen = set()
    keys = []
    for c in companies:
        for rep in c.get("reports", []):
            fin_raw = rep.get("_financial_raw", {})
            fin_ns  = rep.get("_financial_ns", {})
            for k in fin_raw:
                if k not in seen:
                    seen.add(k)
                    keys.append((fin_ns.get(k, ""), k))
    return sorted(keys, key=lambda t: (t[0], t[1]))


def create_fsa_sheet(ws, companies: list):
    ws.title = "Financial Items"

    fin_keys    = _collect_financial_keys(companies)
    col_headers = [f"{_ns_prefix(ns)}:{local}" for ns, local in fin_keys]
    id_headers  = ["CVR", "Company", "Year", "Period end", "Context ID"]
    _header_row(ws, id_headers + col_headers)

    row = 2
    for c in companies:
        cvr  = str(c.get("cvr", ""))
        name = c.get("name", "")

        for rep in c.get("reports", []):
            zebra      = _zebra(row)
            fin_raw    = rep.get("_financial_raw", {})
            context_id = rep.get("_main_context_id", "")

            for col, val in enumerate(
                [cvr, name, rep.get("year", ""), rep.get("period_end", ""), context_id], 1
            ):
                _s(ws.cell(row=row, column=col), value=val, fill=zebra)

            for col, (_, local) in enumerate(fin_keys, 6):
                cell = ws.cell(row=row, column=col)
                if local in fin_raw:
                    _s(cell, value=fin_raw[local], fill=zebra, align="right", fmt=_FMT_INT)
                else:
                    _s(cell, value=_NOT_FOUND, fill=_ORANGE, align="center")

            row += 1

    ws.freeze_panes = "E2"
    _widths(ws, [12, 30, 8, 14, 28] + [32] * len(fin_keys))


# ── Sheet 5: Miscellaneous report items ───────────────────────────────────────

def _collect_misc_keys(companies: list) -> list:
    seen = set()
    keys = []
    for c in companies:
        for rep in c.get("reports", []):
            for k in rep.get("_misc_raw", {}):
                if k not in seen:
                    seen.add(k)
                    keys.append(k)
    return sorted(keys)


def create_misc_sheet(ws, companies: list):
    ws.title = "Miscellaneous report items"

    misc_keys  = _collect_misc_keys(companies)
    id_headers = ["CVR", "Company", "Year", "Period end", "Context ID"]
    _header_row(ws, id_headers + misc_keys)

    row = 2
    for c in companies:
        cvr  = str(c.get("cvr", ""))
        name = c.get("name", "")

        for rep in c.get("reports", []):
            zebra      = _zebra(row)
            misc_raw   = rep.get("_misc_raw", {})
            context_id = rep.get("_main_context_id", "")

            for col, val in enumerate(
                [cvr, name, rep.get("year", ""), rep.get("period_end", ""), context_id], 1
            ):
                _s(ws.cell(row=row, column=col), value=val, fill=zebra)

            for col, key in enumerate(misc_keys, 6):
                cell = ws.cell(row=row, column=col)
                val  = misc_raw.get(key)
                if val is not None:
                    _s(cell, value=val, fill=zebra)
                else:
                    _s(cell, value=_NOT_FOUND, fill=_ORANGE, align="center")

            row += 1

    ws.freeze_panes = "E2"
    _widths(ws, [12, 30, 8, 14, 28] + [40] * len(misc_keys))


# ── Entry point ────────────────────────────────────────────────────────────────

def _create_input_sheet(ws, df):
    """Write the original input DataFrame to a sheet called 'Input'."""
    import pandas as pd
    ws.title = "Input"
    headers = list(df.columns)
    _header_row(ws, headers)
    for row_idx, (_, series) in enumerate(df.iterrows(), 2):
        zebra = _zebra(row_idx)
        for col_idx, val in enumerate(series, 1):
            if pd.isna(val) if not isinstance(val, str) else False:
                val = ""
            _s(ws.cell(row=row_idx, column=col_idx), value=val, fill=zebra)
    ws.freeze_panes = "A2"
    for col in ws.iter_cols(1, ws.max_column):
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)


def create_output(companies: list, extra_input_cols: list, output_file: str, input_df=None):
    """Build the Excel workbook and save it."""
    cur = _determine_currency(companies)

    wb = openpyxl.Workbook()

    create_summary_sheet(wb.active,           companies, extra_input_cols, cur)
    create_company_sheet(wb.create_sheet(),   companies)
    create_financial_sheet(wb.create_sheet(), companies, cur)
    create_fsa_sheet(wb.create_sheet(),       companies)
    create_misc_sheet(wb.create_sheet(),      companies)
    if input_df is not None:
        _create_input_sheet(wb.create_sheet(), input_df)

    wb.save(output_file)
    print(f"Excel gemt: {output_file}")
